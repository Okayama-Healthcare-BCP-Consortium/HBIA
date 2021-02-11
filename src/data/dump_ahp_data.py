import openpyxl
import pandas as pd
import itertools


if __name__ == "__main__":

    df = pd.read_excel('../../data/input/基準・業務リスト.xlsx')

    std_pairs = list(itertools.combinations(df['基準一覧'].dropna(), 2))
    tsk_pairs = list(itertools.combinations(df['業務一覧'].dropna(), 2))

    wb = openpyxl.load_workbook('../../data/raw/AHPアンケートorg.xlsx')
    ws = wb.worksheets[0]

    for i, std_pair in enumerate(std_pairs):
        
        ws[f'A{2+i}'].value = std_pair[0]
        ws[f'K{2+i}'].value = std_pair[1]
        ws[f'F{2+i}'].value = '○'

    for sheet_name in df['基準一覧'].dropna():
        
        ws_copy = wb.copy_worksheet(ws)
        ws_copy.title = sheet_name
        
        for i, tsk_pair in enumerate(tsk_pairs):
            
            ws_copy[f'A{2+i}'].value = tsk_pair[0]
            ws_copy[f'K{2+i}'].value = tsk_pair[1]
            ws_copy[f'F{2+i}'].value = '○'

    wb.save('../../data/input/AHPアンケート.xlsx')

