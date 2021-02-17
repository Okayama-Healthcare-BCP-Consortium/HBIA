import openpyxl
import pandas as pd
import numpy as np
import itertools
import random


def dump_ahp_data(raw_path='./data/raw/AHPアンケートorg.xlsx', out_path='./data/input/AHPアンケート.xlsx'):

    std_pairs = list(itertools.combinations(df['基準一覧'].dropna(), 2))
    tsk_pairs = list(itertools.combinations(df['業務一覧'].dropna(), 2))

    wb = openpyxl.load_workbook(raw_path)
    ws = wb.worksheets[0]

    for i, std_pair in enumerate(std_pairs):
        
        ws[f'A{2+i}'].value = std_pair[0]
        ws[f'K{2+i}'].value = std_pair[1]
        ws[f'F{2+i}'].value = '○'

    for sheet_name in df['基準一覧'].dropna():
        
        ws_copy = wb.copy_worksheet(ws)
        ws_copy.title = sheet_name
        
        for i, tsk_pair in enumerate(tsk_pairs):
            
            ws_copy[f'A{2+i}'].value = tsk_pair[0]
            ws_copy[f'K{2+i}'].value = tsk_pair[1]
            ws_copy[f'F{2+i}'].value = '○'

    wb.save(out_path)


def dump_reqfil_data(raw_path='./data/raw/必要物フィルタのための入力データorg.xlsx', out_path='./data/input/必要物フィルタのための入力データ.xlsx'):

    samp_nums = np.arange(1, 3+1)
    scores = np.arange(0, 2+1)

    wb = openpyxl.load_workbook(raw_path)

    ws = wb.worksheets[0]

    for i, req in enumerate(df['必要物一覧'].dropna()):     

        ws[f'A{2+i}'].value = req
        ws[f'B{2+i}'].value = random.choice(scores)

    ws = wb.worksheets[1]

    for i, tsk in enumerate(df['業務一覧'].dropna()): 

        samp_num = random.choice(samp_nums)
        reqs = random.sample(list(df['必要物一覧'].dropna()), samp_num)

        ws[f'A{2+i}'].value = tsk
        ws[f'B{2+i}'].value = ','.join(reqs)

    wb.save(out_path)


if __name__ == "__main__":

    path = './data/input/基準・業務・必要物リスト.xlsx'
    df = pd.read_excel(path)

    dump_ahp_data()
    dump_reqfil_data()
